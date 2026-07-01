"""
extrator_ocr.py - Módulo de extração OCR para listas de presença.

Converte PDF em imagens, executa Tesseract OCR e exporta para CSV/XLSX.
"""
from __future__ import annotations

import csv
import os
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence


def _find_tesseract() -> str:
    """Localiza o executável do Tesseract (Windows e Linux)."""
    if shutil.which("tesseract"):
        return "tesseract"
    # Caminhos comuns no Windows
    candidates = [
        Path(os.environ.get("PROGRAMFILES", "")) / "Tesseract-OCR" / "tesseract.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", "")) / "Tesseract-OCR" / "tesseract.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tesseract.exe",
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return ""


TESSERACT_CMD = _find_tesseract()


@dataclass
class ExtractedRow:
    source: str
    page: int
    row_number: int
    columns: list[str] = field(default_factory=list)


def count_pdf_pages(pdf_path: Path) -> int:
    """Conta o numero de paginas de um PDF."""
    try:
        import fitz  # PyMuPDF

        with fitz.open(str(pdf_path)) as doc:
            return len(doc)
    except ImportError:
        pass

    try:
        from pdf2image import pdfinfo_from_path

        info = pdfinfo_from_path(str(pdf_path))
        return int(info.get("Pages", 0))
    except ImportError:
        pass

    raise RuntimeError(
        "Nenhuma biblioteca de conversao PDF disponivel. "
        "Instale PyMuPDF (pip install pymupdf) ou pdf2image + poppler."
    )


def render_pdf_page(
    pdf_path: Path,
    page_number: int,
    output_path: Path,
    dpi: int = 300,
    image_format: str = "png",
    jpeg_quality: int = 85,
) -> Path:
    """Renderiza uma pagina especifica do PDF em alta qualidade."""
    normalized_format = image_format.strip().lower()
    if normalized_format not in {"png", "jpg", "jpeg"}:
        normalized_format = "png"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        import fitz  # PyMuPDF

        with fitz.open(str(pdf_path)) as doc:
            page = doc[page_number - 1]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            if normalized_format == "png":
                pix.save(str(output_path))
            else:
                from PIL import Image

                image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                image.save(str(output_path), format="JPEG", quality=jpeg_quality, optimize=True)
        return output_path
    except ImportError:
        pass

    try:
        from pdf2image import convert_from_path

        pil_images = convert_from_path(
            str(pdf_path),
            dpi=dpi,
            first_page=page_number,
            last_page=page_number,
            fmt="jpg" if normalized_format in {"jpg", "jpeg"} else "png",
        )
        image = pil_images[0]
        if normalized_format == "png":
            image.save(str(output_path))
        else:
            image.convert("RGB").save(str(output_path), format="JPEG", quality=jpeg_quality, optimize=True)
        return output_path
    except ImportError:
        pass

    raise RuntimeError(
        "Nenhuma biblioteca de conversao PDF disponivel. "
        "Instale PyMuPDF (pip install pymupdf) ou pdf2image + poppler."
    )


def pdf_to_images(
    pdf_path: Path,
    output_dir: Path,
    dpi: int = 300,
    image_format: str = "png",
    jpeg_quality: int = 85,
) -> list[Path]:
    """Converte PDF em imagens usando pdf2image (poppler) ou fallback com PyMuPDF."""
    images: list[Path] = []
    normalized_format = image_format.strip().lower()
    if normalized_format not in {"png", "jpg", "jpeg"}:
        normalized_format = "png"
    extension = "jpg" if normalized_format in {"jpg", "jpeg"} else "png"

    # Tenta PyMuPDF (fitz) primeiro - mais confiável no Windows
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(str(pdf_path))
        for page_num in range(len(doc)):
            page = doc[page_num]
            # Calcula zoom para atingir o DPI desejado (padrão PDF = 72 DPI)
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_path = output_dir / f"page_{page_num + 1:04d}.{extension}"
            if extension == "png":
                pix.save(str(img_path))
            else:
                from PIL import Image

                image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                image.save(str(img_path), format="JPEG", quality=jpeg_quality, optimize=True)
            images.append(img_path)
        doc.close()
        return images
    except ImportError:
        pass

    # Fallback: pdf2image (requer poppler instalado)
    try:
        from pdf2image import convert_from_path
        pil_images = convert_from_path(str(pdf_path), dpi=dpi, output_folder=str(output_dir), fmt=extension)
        for idx, img in enumerate(pil_images, start=1):
            img_path = output_dir / f"page_{idx:04d}.{extension}"
            if extension == "png":
                img.save(str(img_path))
            else:
                rgb_img = img.convert("RGB")
                rgb_img.save(str(img_path), format="JPEG", quality=jpeg_quality, optimize=True)
            images.append(img_path)
        return images
    except ImportError:
        pass

    raise RuntimeError(
        "Nenhuma biblioteca de conversão PDF disponível. "
        "Instale PyMuPDF (pip install pymupdf) ou pdf2image + poppler."
    )


def run_tesseract(image_path: Path, lang: str = "por+eng", psm: str = "6") -> str:
    """Executa Tesseract OCR em uma imagem e retorna o texto extraído."""
    if not TESSERACT_CMD:
        raise RuntimeError(
            "Tesseract OCR não encontrado. "
            "Instale: https://github.com/UB-Mannheim/tesseract/wiki"
        )
    cmd = [TESSERACT_CMD, str(image_path), "stdout", "-l", lang, "--psm", psm]
    result = subprocess.run(cmd, capture_output=True, timeout=120, encoding="utf-8", errors="replace")
    if result.returncode != 0:
        raise RuntimeError(f"Tesseract falhou: {result.stderr[:500]}")
    return result.stdout


def extract_rows(
    files: list[Path],
    lang: str = "por+eng",
    psm: str = "6",
    dpi: int = 300,
    mode: str = "table",
) -> list[ExtractedRow]:
    """Extrai linhas de texto de PDFs/imagens via OCR."""
    rows: list[ExtractedRow] = []

    with tempfile.TemporaryDirectory(prefix="ocr_extract_") as temp:
        work_dir = Path(temp)

        for file_path in files:
            if file_path.suffix.lower() == ".pdf":
                image_paths = pdf_to_images(file_path, work_dir, dpi)
            else:
                image_paths = [file_path]

            for page_num, img_path in enumerate(image_paths, start=1):
                text = run_tesseract(img_path, lang, psm)

                if mode == "table":
                    # Modo tabela: tenta separar por tabulações ou múltiplos espaços
                    for line_num, line in enumerate(text.splitlines(), start=1):
                        line = line.strip()
                        if not line:
                            continue
                        # Separa por tab ou 3+ espaços
                        import re
                        columns = re.split(r'\t|  {2,}', line)
                        columns = [col.strip() for col in columns if col.strip()]
                        if columns:
                            rows.append(ExtractedRow(str(file_path), page_num, line_num, columns))
                else:
                    # Modo linhas: cada linha é uma coluna única
                    for line_num, line in enumerate(text.splitlines(), start=1):
                        line = line.strip()
                        if line:
                            rows.append(ExtractedRow(str(file_path), page_num, line_num, [line]))

    return rows


def write_output(
    rows: list[ExtractedRow],
    output_path: Path,
    headers: Sequence[str] | None = None,
) -> None:
    """Escreve os dados extraídos em CSV ou XLSX com formatação condicional."""
    suffix = output_path.suffix.lower()

    if suffix == ".csv":
        _write_csv(rows, output_path, headers)
    elif suffix == ".xlsx":
        _write_xlsx(rows, output_path, headers)
    else:
        # Default para xlsx
        output_path = output_path.with_suffix(".xlsx")
        _write_xlsx(rows, output_path, headers)


def _output_row_values(row: ExtractedRow, headers: Sequence[str] | None = None) -> list[str]:
    values = list(row.columns)
    if headers:
        return values[:len(headers)]
    return values


def _write_csv(
    rows: list[ExtractedRow],
    output_path: Path,
    headers: Sequence[str] | None = None,
) -> None:
    """Exporta para CSV."""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")

        if headers:
            writer.writerow(headers)
        else:
            # Header genérico baseado no máximo de colunas
            max_cols = max((len(r.columns) for r in rows), default=0)
            writer.writerow(["Arquivo", "Página", "Linha"] + [f"Coluna_{i+1}" for i in range(max_cols)])

        for row in rows:
            writer.writerow([row.source, row.page, row.row_number] + _output_row_values(row, headers))


def _write_xlsx(
    rows: list[ExtractedRow],
    output_path: Path,
    headers: Sequence[str] | None = None,
) -> None:
    """Exporta para XLSX com formatação condicional (verde=presente, amarelo=ausente)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        # Fallback para CSV se openpyxl não estiver disponível
        csv_path = output_path.with_suffix(".csv")
        _write_csv(rows, csv_path, headers)
        # Renomeia para .xlsx (será CSV na verdade, mas funciona como fallback)
        import shutil
        shutil.copy2(csv_path, output_path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Presença"

    # Estilos
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    presente_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ausente_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    if headers:
        header_row = list(headers)
    else:
        max_cols = max((len(r.columns) for r in rows), default=0)
        header_row = ["Arquivo", "Página", "Linha"] + [f"Coluna_{i+1}" for i in range(max_cols)]

    for col_idx, header in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Índice da coluna de presença para formatação condicional
    assinou_col_idx = None
    if headers:
        for idx, h in enumerate(headers):
            if any(kw in h.lower() for kw in ("assinou", "presente", "assinatura")):
                assinou_col_idx = idx
                break

    # Dados
    for row_idx, row in enumerate(rows, start=2):
        data = _output_row_values(row, headers)
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Formatação condicional na coluna de presença
            if assinou_col_idx is not None and col_idx == assinou_col_idx + 1:
                if value.lower().strip() in ("sim", "presente", "yes", "s"):
                    cell.fill = presente_fill
                    cell.font = Font(bold=True, color="006100")
                elif value.lower().strip() in ("nao", "não", "ausente", "no", "n"):
                    cell.fill = ausente_fill
                    cell.font = Font(bold=True, color="9C5700")

    # Auto-ajuste de largura
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(50, max(12, max_length + 2))

    # Congela header
    ws.freeze_panes = "A2"

    wb.save(output_path)
