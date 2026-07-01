from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from extrator_ocr import ExtractedRow


class ProcessingProfileTests(unittest.TestCase):
    def test_selects_small_profile_for_small_pdf(self) -> None:
        from attendance_pipeline import select_processing_profile

        profile = select_processing_profile(Path("small.pdf"), page_count=4, file_size_bytes=6 * 1024 * 1024)

        self.assertEqual("small", profile.name)
        self.assertEqual(180, profile.dpi)
        self.assertEqual("jpeg", profile.image_format)
        self.assertEqual(82, profile.jpeg_quality)
        self.assertEqual(4, profile.max_concurrency)
        self.assertEqual(8, profile.min_rows_per_page)

    def test_selects_medium_profile_by_pages_and_size(self) -> None:
        from attendance_pipeline import select_processing_profile

        profile = select_processing_profile(Path("medium.pdf"), page_count=12, file_size_bytes=15 * 1024 * 1024)

        self.assertEqual("medium", profile.name)
        self.assertEqual(130, profile.dpi)
        self.assertEqual(75, profile.jpeg_quality)
        self.assertEqual(6, profile.max_concurrency)
        self.assertEqual(8, profile.min_rows_per_page)

    def test_selects_large_profile_when_document_exceeds_medium_limits(self) -> None:
        from attendance_pipeline import select_processing_profile

        profile = select_processing_profile(Path("large.pdf"), page_count=20, file_size_bytes=16 * 1024 * 1024)

        self.assertEqual("large", profile.name)
        self.assertEqual(120, profile.dpi)
        self.assertEqual(68, profile.jpeg_quality)
        self.assertEqual(6, profile.max_concurrency)
        self.assertEqual(6, profile.min_rows_per_page)

    def test_allows_min_rows_override_via_env(self) -> None:
        from attendance_pipeline import select_processing_profile

        with mock.patch.dict("os.environ", {"OCR_GEMINI_MIN_ROWS_PER_PAGE": "11"}, clear=False):
            profile = select_processing_profile(Path("small.pdf"), page_count=2, file_size_bytes=1 * 1024 * 1024)

        self.assertEqual(11, profile.min_rows_per_page)

    def test_stable_production_mode_locks_profile_defaults(self) -> None:
        from attendance_pipeline import select_processing_profile

        with mock.patch.dict(
            "os.environ",
            {
                "OCR_STABLE_PRODUCTION_MODE": "true",
                "OCR_GEMINI_MEDIUM_DPI": "123",
                "OCR_GEMINI_MAX_CONCURRENCY": "9",
            },
            clear=False,
        ):
            profile = select_processing_profile(Path("medium.pdf"), page_count=10, file_size_bytes=8 * 1024 * 1024)

        self.assertEqual("medium", profile.name)
        self.assertEqual(160, profile.dpi)
        self.assertEqual(4, profile.max_concurrency)


class WarmupCacheTests(unittest.TestCase):
    def test_reuses_gemini_warmup_within_ttl(self) -> None:
        from attendance_pipeline import maybe_warmup_gemini_runtime

        with mock.patch("attendance_pipeline._GEMINI_WARMUP_CACHE", None), \
            mock.patch("attendance_pipeline._GEMINI_WARMUP_CACHE_TS", 0.0), \
            mock.patch("attendance_pipeline._gemini_warmup_ttl_seconds", return_value=600), \
            mock.patch("attendance_pipeline.warmup_gemini_runtime", return_value={"token_ready": True, "fast_model_available": False, "elapsed_ms": 7, "error": ""}) as warmup_mock:
            first = maybe_warmup_gemini_runtime(timeout_seconds=8)
            second = maybe_warmup_gemini_runtime(timeout_seconds=8)

        self.assertFalse(first.get("cached"))
        self.assertTrue(second.get("cached"))
        self.assertEqual(1, warmup_mock.call_count)


class FinalizationTests(unittest.TestCase):
    def test_finalizes_results_in_page_order_and_inherits_headers(self) -> None:
        from attendance_pipeline import AttendancePageResult, finalize_page_results

        page_two = AttendancePageResult(
            page_number=2,
            rows=[
                {
                    "nome": "Aluno Dois",
                    "matutino_status": "Presente",
                    "matutino_texto": "Rubrica",
                    "vespertino_status": "Ausente",
                    "vespertino_texto": "",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                }
            ],
            header={"modulo": "", "curso": "", "turma": "", "data": ""},
            processor_used="gemini",
            timings_ms={"gemini_ms": 10},
        )
        page_one = AttendancePageResult(
            page_number=1,
            rows=[
                {
                    "nome": "Aluno Um",
                    "matutino_status": "Presente",
                    "matutino_texto": "Sim",
                    "vespertino_status": "Ausente",
                    "vespertino_texto": "",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                }
            ],
            header={"modulo": "Modulo I", "curso": "Curso X", "turma": "Turma 7", "data": "01/01/2026"},
            processor_used="gemini",
            timings_ms={"gemini_ms": 9},
        )

        rows = finalize_page_results(Path("arquivo.pdf"), [page_two, page_one])

        self.assertEqual(6, len(rows))
        self.assertEqual("Aluno Um", rows[0].columns[4])
        self.assertEqual("MÓDULO I", rows[0].columns[0])
        self.assertEqual("CURSO DE FORMAÇÃO EM Curso X", rows[0].columns[1])
        self.assertEqual("01/01/2026", rows[0].columns[3])
        self.assertEqual("Aluno Dois", rows[3].columns[4])
        self.assertEqual("CURSO DE FORMAÇÃO EM Curso X", rows[3].columns[1])
        self.assertEqual("Turma 7", rows[3].columns[2])
        self.assertEqual("01/01/2026", rows[3].columns[3])


class OutputTests(unittest.TestCase):
    def test_write_output_trims_extra_columns_to_header_count(self) -> None:
        from openpyxl import load_workbook

        from extrator_ocr import ExtractedRow, write_output

        headers = [f"Coluna {index}" for index in range(1, 9)]
        row = ExtractedRow(
            source="arquivo.pdf",
            page=1,
            row_number=1,
            columns=["a", "b", "c", "d", "e", "f", "g", "h", "i"],
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "saida.xlsx"
            write_output([row], output_path, headers)
            ws = load_workbook(output_path).active

        self.assertEqual(8, ws.max_column)
        self.assertEqual(headers, [ws.cell(row=1, column=index).value for index in range(1, 9)])
        self.assertEqual(["a", "b", "c", "d", "e", "f", "g", "h"], [ws.cell(row=2, column=index).value for index in range(1, 9)])


class NameDedupTests(unittest.TestCase):
    def test_deduplicates_common_ocr_name_variants(self) -> None:
        from attendance_pipeline import _deduplicate_names

        rows = [
            ExtractedRow("arquivo.pdf", 1, 1, ["M", "C", "T", "01/01/2026", "Alciele Fernandes Peixoto", "Matutino", "Ausente", "nao_assinado", ""]),
            ExtractedRow("arquivo.pdf", 1, 2, ["M", "C", "T", "01/01/2026", "Alcicle Fernandes Peixoto", "Vespertino", "Ausente", "nao_assinado", ""]),
            ExtractedRow("arquivo.pdf", 1, 3, ["M", "C", "T", "01/01/2026", "Larissa Miranda dos Santos", "Noturno", "Ausente", "nao_assinado", ""]),
            ExtractedRow("arquivo.pdf", 1, 4, ["M", "C", "T", "01/01/2026", "Loarissa Miranda dos Santos", "Matutino", "Ausente", "nao_assinado", ""]),
        ]

        deduped = _deduplicate_names(rows)
        names = [row.columns[4] for row in deduped]

        self.assertIn("Alciele Fernandes Peixoto", names)
        self.assertNotIn("Alcicle Fernandes Peixoto", names)
        self.assertIn("Larissa Miranda dos Santos", names)
        self.assertNotIn("Loarissa Miranda dos Santos", names)


class GuardTests(unittest.TestCase):
    def test_detects_partial_name_copy_as_suspicious(self) -> None:
        from attendance_pipeline import _looks_like_same_as_printed_name

        self.assertTrue(_looks_like_same_as_printed_name("Alexandre Paniago de Oliveira", "Alexandre"))
        self.assertTrue(_looks_like_same_as_printed_name("Alessandra M. Cardoso", "Alessandra"))
        self.assertTrue(_looks_like_same_as_printed_name("Adriano Matias Quiste", "A. Quiste"))
        self.assertFalse(_looks_like_same_as_printed_name("Bianca Nunes de Queiroz", "Presente"))

    def test_visual_absence_guard_aligns_rows_after_header_offset(self) -> None:
        from attendance_pipeline import _apply_visual_absence_guard

        class DummyImage:
            width = 1200

            def convert(self, _mode: str) -> "DummyImage":
                return self

        class DummyInk:
            def __init__(self, signed: bool) -> None:
                self.signed = signed
                self.text = ""

        rows = [
            {
                "nome": f"Aluno {index}",
                "matutino_status": "Ausente",
                "matutino_texto": "",
                "vespertino_status": "Ausente",
                "vespertino_texto": "",
                "noturno_status": "Ausente",
                "noturno_texto": "",
            }
            for index in range(1, 24)
        ]
        rows[4]["nome"] = "Bianca Nunes de Queiroz"
        rows[4]["matutino_status"] = "Presente"
        rows[4]["matutino_texto"] = "Sim"
        rows[4]["vespertino_status"] = "Presente"
        rows[4]["vespertino_texto"] = "Sim"
        rows[4]["noturno_status"] = "Presente"
        rows[4]["noturno_texto"] = "Sim"

        row_intervals = [(index * 10, (index * 10) + 8) for index in range(25)]
        expected_top = row_intervals[6][0]

        def fake_crop(_image, _left, top, _right, _bottom, _margin):
            return {"top": top}

        def fake_analyze(cell, _lang, include_text=True):
            del include_text
            return DummyInk(signed=cell["top"] != expected_top)

        with mock.patch("PIL.Image.open", return_value=DummyImage()), \
            mock.patch("assinatura_lista._auto_rotate", side_effect=lambda image: image), \
            mock.patch("assinatura_lista.detect_table_grid", return_value=mock.Mock(horizontal=list(range(25)), vertical=[0, 400, 800, 1000])), \
            mock.patch("assinatura_lista.data_row_intervals", return_value=row_intervals), \
            mock.patch("assinatura_lista.signature_column_intervals", return_value=[(0, 10), (10, 20), (20, 30)]), \
            mock.patch("assinatura_lista.crop_with_margin", side_effect=fake_crop), \
            mock.patch("assinatura_lista.analyze_cell_ink", side_effect=fake_analyze):
            fixed_rows = _apply_visual_absence_guard(Path("page.jpg"), rows, "pt")

        self.assertEqual("Ausente", fixed_rows[4]["matutino_status"])
        self.assertEqual("Ausente", fixed_rows[4]["vespertino_status"])
        self.assertEqual("Ausente", fixed_rows[4]["noturno_status"])


class OrchestratorTests(unittest.TestCase):
    def test_processes_pages_with_selective_fallback(self) -> None:
        from attendance_pipeline import ProcessingProfile, process_attendance_list

        profile = ProcessingProfile(
            name="medium",
            dpi=160,
            image_format="jpeg",
            jpeg_quality=75,
            max_concurrency=3,
            min_rows_per_page=8,
            fallback_mode="documentai",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            pdf_path = base / "documento.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 fake")
            output_path = base / "saida.xlsx"
            page1 = base / "page_0001.jpg"
            page2 = base / "page_0002.jpg"
            page1.write_bytes(b"page1")
            page2.write_bytes(b"page2")

            gemini_page_one = {
                "header": {"modulo": "Modulo II", "curso": "Curso Y", "turma": "Turma 9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": f"Aluno {index}",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                    for index in range(1, 9)
                ],
            }
            fallback_rows = [
                ExtractedRow(
                    source=str(page2),
                    page=1,
                    row_number=1,
                    columns=[
                        "Modulo II",
                        "Curso Y",
                        "Turma 9",
                        "02/02/2026",
                        "Fallback Dois",
                        "Matutino",
                        "Presente",
                        "rubrica",
                        "Ass",
                    ],
                )
            ]

            def gemini_side_effect(image_path: Path, lang: str = "pt", **_kwargs) -> dict[str, object]:
                del lang
                if image_path == page1:
                    return gemini_page_one
                if image_path == page2:
                    raise RuntimeError("falhou")
                raise AssertionError(f"unexpected image path: {image_path}")

            with mock.patch("attendance_pipeline.inspect_pdf_document", return_value=(2, "application/pdf")), \
                mock.patch("attendance_pipeline.select_processing_profile", return_value=profile), \
                mock.patch("attendance_pipeline.render_pdf_for_profile", return_value=[page1, page2]), \
                mock.patch("attendance_pipeline.maybe_warmup_gemini_runtime", return_value={"token_ready": True, "fast_model_available": False, "elapsed_ms": 1, "error": "", "cached": False}), \
                mock.patch("attendance_pipeline._fast_model_first_pass_enabled", return_value=False), \
                mock.patch("attendance_pipeline._documentai_fallback_available", return_value=(True, "ok")), \
                mock.patch("attendance_pipeline.process_pages_with_gemini", side_effect=RuntimeError("batch indisponivel")), \
                mock.patch("attendance_pipeline.process_page_with_gemini", side_effect=gemini_side_effect), \
                mock.patch("attendance_pipeline.process_page_with_documentai", return_value=fallback_rows) as docai_mock, \
                mock.patch("attendance_pipeline.write_output") as write_output_mock:
                row_count = process_attendance_list(pdf_path, output_path, lang="pt")

        self.assertEqual(27, row_count)
        self.assertEqual(1, docai_mock.call_count)
        write_output_mock.assert_called_once()
        written_rows = write_output_mock.call_args.args[0]
        self.assertEqual("Aluno 1", written_rows[0].columns[4])
        self.assertEqual("Fallback Dois", written_rows[-3].columns[4])
        self.assertEqual("CURSO DE FORMAÇÃO EM Curso Y", written_rows[-3].columns[1])

    def test_retries_low_quality_gemini_page_in_high_quality_before_fallback(self) -> None:
        from attendance_pipeline import ProcessingProfile, process_attendance_list

        profile = ProcessingProfile(
            name="medium",
            dpi=160,
            image_format="jpeg",
            jpeg_quality=75,
            max_concurrency=3,
            min_rows_per_page=8,
            fallback_mode="documentai",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            pdf_path = base / "documento.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 fake")
            output_path = base / "saida.xlsx"
            page1 = base / "page_0001.jpg"
            hq_page1 = base / "page_0001.png"
            page1.write_bytes(b"page1")
            hq_page1.write_bytes(b"page1-hq")

            low_quality_result = {
                "header": {"modulo": "Modulo II", "curso": "Curso Y", "turma": "9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": "Aluno 1",
                        "matutino_status": "Ausente",
                        "matutino_texto": "",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                ],
            }
            high_quality_result = {
                "header": {"modulo": "MÓDULO II", "curso": "CURSO DE FORMAÇÃO EM Curso Y", "turma": "Turma 9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": f"Aluno {index}",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Presente",
                        "vespertino_texto": "Ass",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                    for index in range(1, 10)
                ],
            }

            with mock.patch("attendance_pipeline.inspect_pdf_document", return_value=(1, "application/pdf")), \
                mock.patch("attendance_pipeline.select_processing_profile", return_value=profile), \
                mock.patch("attendance_pipeline.render_pdf_for_profile", return_value=[page1]), \
                mock.patch("attendance_pipeline.maybe_warmup_gemini_runtime", return_value={"token_ready": True, "fast_model_available": False, "elapsed_ms": 1, "error": "", "cached": False}), \
                mock.patch("attendance_pipeline._fast_model_first_pass_enabled", return_value=False), \
                mock.patch("attendance_pipeline._high_quality_retry_enabled", return_value=True), \
                mock.patch("attendance_pipeline._allow_low_row_gemini_result", return_value=True), \
                mock.patch("attendance_pipeline.render_pdf_page", return_value=hq_page1) as hq_render_mock, \
                mock.patch("attendance_pipeline.process_page_with_gemini", side_effect=[low_quality_result, high_quality_result]), \
                mock.patch("attendance_pipeline.process_page_with_documentai") as docai_mock, \
                mock.patch("attendance_pipeline.write_output") as write_output_mock:
                row_count = process_attendance_list(pdf_path, output_path, lang="pt")

        self.assertEqual(27, row_count)
        hq_render_mock.assert_called_once()
        docai_mock.assert_not_called()
        written_rows = write_output_mock.call_args.args[0]
        self.assertEqual("Aluno 1", written_rows[0].columns[4])
        self.assertEqual("MÓDULO II", written_rows[0].columns[0])
        self.assertEqual("CURSO DE FORMAÇÃO EM Curso Y", written_rows[0].columns[1])

    def test_merges_low_and_high_quality_gemini_rows(self) -> None:
        from attendance_pipeline import ProcessingProfile, process_attendance_list

        profile = ProcessingProfile(
            name="medium",
            dpi=160,
            image_format="jpeg",
            jpeg_quality=75,
            max_concurrency=3,
            min_rows_per_page=8,
            fallback_mode="documentai",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            pdf_path = base / "documento.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 fake")
            output_path = base / "saida.xlsx"
            page1 = base / "page_0001.jpg"
            hq_page1 = base / "page_0001.png"
            page1.write_bytes(b"page1")
            hq_page1.write_bytes(b"page1-hq")

            low_quality_result = {
                "header": {"modulo": "Modulo II", "curso": "Curso Y", "turma": "9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": "Aluno 1",
                        "matutino_status": "Ausente",
                        "matutino_texto": "",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                ],
            }
            high_quality_result = {
                "header": {"modulo": "MÓDULO II", "curso": "CURSO DE FORMAÇÃO EM Curso Y", "turma": "Turma 9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": "Aluno 1",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    },
                    {
                        "nome": "Aluno 2",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    },
                ],
            }

            with mock.patch("attendance_pipeline.inspect_pdf_document", return_value=(1, "application/pdf")), \
                mock.patch("attendance_pipeline.select_processing_profile", return_value=profile), \
                mock.patch("attendance_pipeline.render_pdf_for_profile", return_value=[page1]), \
                mock.patch("attendance_pipeline.maybe_warmup_gemini_runtime", return_value={"token_ready": True, "fast_model_available": False, "elapsed_ms": 1, "error": "", "cached": False}), \
                mock.patch("attendance_pipeline._fast_model_first_pass_enabled", return_value=False), \
                mock.patch("attendance_pipeline._high_quality_retry_enabled", return_value=True), \
                mock.patch("attendance_pipeline._allow_low_row_gemini_result", return_value=True), \
                mock.patch("attendance_pipeline.render_pdf_page", return_value=hq_page1), \
                mock.patch("attendance_pipeline.process_page_with_gemini", side_effect=[low_quality_result, high_quality_result]), \
                mock.patch("attendance_pipeline.process_page_with_documentai") as docai_mock, \
                mock.patch("attendance_pipeline.write_output") as write_output_mock:
                row_count = process_attendance_list(pdf_path, output_path, lang="pt")

        self.assertEqual(6, row_count)
        docai_mock.assert_not_called()
        written_rows = write_output_mock.call_args.args[0]
        self.assertEqual("Aluno 1", written_rows[0].columns[4])
        self.assertEqual("Presente", written_rows[0].columns[6])
        self.assertEqual("Aluno 2", written_rows[3].columns[4])

    def test_refines_previous_page_when_next_page_is_sparse(self) -> None:
        from attendance_pipeline import ProcessingProfile, process_attendance_list

        profile = ProcessingProfile(
            name="medium",
            dpi=160,
            image_format="jpeg",
            jpeg_quality=75,
            max_concurrency=3,
            min_rows_per_page=1,
            fallback_mode="documentai",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            pdf_path = base / "documento.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 fake")
            output_path = base / "saida.xlsx"
            page1 = base / "page_0001.jpg"
            page2 = base / "page_0002.jpg"
            page1.write_bytes(b"page1")
            page2.write_bytes(b"page2")

            page_one_low = {
                "header": {"modulo": "Modulo II", "curso": "Curso Y", "turma": "9", "data": "02/02/2026"},
                "rows": [
                    {
                        "nome": "Aluno 1",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                ] + [
                    {
                        "nome": f"Aluno {index}",
                        "matutino_status": "Ausente",
                        "matutino_texto": "",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                    for index in range(2, 22)
                ],
            }
            page_two_sparse = {
                "header": {},
                "rows": [
                    {
                        "nome": "Aluno 22",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    }
                ],
            }
            page_one_hq = {
                "header": {"modulo": "MÓDULO II", "curso": "CURSO DE FORMAÇÃO EM Curso Y", "turma": "Turma 9", "data": "02/02/2026"},
                "rows": page_one_low["rows"] + [
                    {
                        "nome": "Aluno 23",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    },
                    {
                        "nome": "Aluno 24",
                        "matutino_status": "Presente",
                        "matutino_texto": "Sim",
                        "vespertino_status": "Ausente",
                        "vespertino_texto": "",
                        "noturno_status": "Ausente",
                        "noturno_texto": "",
                    },
                ],
            }

            def gemini_side_effect(image_path: Path, lang: str = "pt", **_kwargs) -> dict[str, object]:
                del lang
                if image_path == page1:
                    return page_one_low
                if image_path == page2:
                    return page_two_sparse
                raise AssertionError(f"unexpected image path: {image_path}")

            def refine_side_effect(_source_path: Path, page_result, _lang: str):
                from attendance_pipeline import AttendancePageResult

                if page_result.page_number != 1:
                    return page_result
                return AttendancePageResult(
                    page_number=1,
                    rows=page_one_hq["rows"],
                    header=page_one_hq["header"],
                    processor_used="gemini_hq",
                    timings_ms={"gemini_ms": 0},
                )

            with mock.patch("attendance_pipeline.inspect_pdf_document", return_value=(2, "application/pdf")), \
                mock.patch("attendance_pipeline.select_processing_profile", return_value=profile), \
                mock.patch("attendance_pipeline.render_pdf_for_profile", return_value=[page1, page2]), \
                mock.patch("attendance_pipeline.maybe_warmup_gemini_runtime", return_value={"token_ready": True, "fast_model_available": False, "elapsed_ms": 1, "error": "", "cached": False}), \
                mock.patch("attendance_pipeline._fast_model_first_pass_enabled", return_value=False), \
                mock.patch("attendance_pipeline._refine_previous_page_enabled", return_value=True), \
                mock.patch("attendance_pipeline._smart_refine_enabled", return_value=False), \
                mock.patch("attendance_pipeline.process_pages_with_gemini", side_effect=RuntimeError("batch indisponivel")), \
                mock.patch("attendance_pipeline.process_page_with_gemini", side_effect=gemini_side_effect), \
                mock.patch("attendance_pipeline._refine_result_with_high_quality", side_effect=refine_side_effect) as refine_mock, \
                mock.patch("attendance_pipeline.process_page_with_documentai") as docai_mock, \
                mock.patch("attendance_pipeline.write_output") as write_output_mock:
                row_count = process_attendance_list(pdf_path, output_path, lang="pt")

        self.assertEqual(72, row_count)
        refine_mock.assert_called_once()
        docai_mock.assert_not_called()
        written_rows = write_output_mock.call_args.args[0]
        self.assertEqual("Aluno 24", written_rows[-6].columns[4])

    def test_retries_suspicious_prefetched_batch_page_with_single_page_gemini(self) -> None:
        from attendance_pipeline import ProcessingProfile, _process_single_page

        profile = ProcessingProfile(
            name="medium",
            dpi=160,
            image_format="jpeg",
            jpeg_quality=75,
            max_concurrency=3,
            min_rows_per_page=8,
            fallback_mode="documentai",
        )

        prefetched_result = {
            "header": {"modulo": "Modulo II", "curso": "Curso Y", "turma": "Turma 9", "data": "02/02/2026"},
            "rows": [
                {
                    "nome": "Alessandra M. Cardoso",
                    "matutino_status": "Presente",
                    "matutino_texto": "Alessandra",
                    "vespertino_status": "Presente",
                    "vespertino_texto": "Alessandra",
                    "noturno_status": "Presente",
                    "noturno_texto": "Alessandra",
                },
                {
                    "nome": "Bianca Nunes de Queiroz",
                    "matutino_status": "Presente",
                    "matutino_texto": "Sim",
                    "vespertino_status": "Presente",
                    "vespertino_texto": "Sim",
                    "noturno_status": "Presente",
                    "noturno_texto": "Sim",
                },
            ] + [
                {
                    "nome": f"Aluno {index}",
                    "matutino_status": "Ausente",
                    "matutino_texto": "",
                    "vespertino_status": "Ausente",
                    "vespertino_texto": "",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                }
                for index in range(3, 11)
            ],
        }
        clean_single_page = {
            "header": prefetched_result["header"],
            "rows": [
                {
                    "nome": "Alessandra M. Cardoso",
                    "matutino_status": "Presente",
                    "matutino_texto": "Assinatura",
                    "vespertino_status": "Presente",
                    "vespertino_texto": "Assinatura",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                },
                {
                    "nome": "Bianca Nunes de Queiroz",
                    "matutino_status": "Ausente",
                    "matutino_texto": "",
                    "vespertino_status": "Ausente",
                    "vespertino_texto": "",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                },
            ] + [
                {
                    "nome": f"Aluno {index}",
                    "matutino_status": "Ausente",
                    "matutino_texto": "",
                    "vespertino_status": "Ausente",
                    "vespertino_texto": "",
                    "noturno_status": "Ausente",
                    "noturno_texto": "",
                }
                for index in range(3, 11)
            ],
        }

        with mock.patch("attendance_pipeline._use_gemini", return_value=True), \
            mock.patch("attendance_pipeline._use_documentai", return_value=False), \
            mock.patch("attendance_pipeline._fast_model_first_pass_enabled", return_value=False), \
            mock.patch("attendance_pipeline._crop_remote_image_if_possible", side_effect=lambda path: path), \
            mock.patch("attendance_pipeline._apply_visual_absence_guard", side_effect=lambda _path, rows, _lang: rows), \
            mock.patch("attendance_pipeline._apply_legacy_signature_guard", side_effect=lambda _image, _source, _page, rows, _lang: rows), \
            mock.patch("attendance_pipeline.process_page_with_gemini", return_value=clean_single_page) as single_page_mock:
            result = _process_single_page(
                Path("arquivo.pdf"),
                3,
                Path("page_0003.jpg"),
                profile,
                "pt",
                prefetched_gemini_result=prefetched_result,
                prefetched_gemini_ms=100,
                expected_name_count_override=10,
            )

        self.assertEqual(1, single_page_mock.call_count)
        self.assertEqual("Assinatura", result.rows[0]["matutino_texto"])
        self.assertEqual("Ausente", result.rows[1]["matutino_status"])


if __name__ == "__main__":
    unittest.main()
